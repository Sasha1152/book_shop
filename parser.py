from openpyxl import load_workbook

workbook = load_workbook('./booklist.xlsx')

print(workbook.sheetnames)

sheet = workbook['Sheet1']
print(sheet.title)

book = {}
author = {}

book['title'] = sheet['a3'].value
author['first_name'] = sheet['b3'].value
author['last_name'] = sheet['c3'].value
author['biography'] = sheet['d3'].value
book['description'] = sheet['e3'].value
book['year'] = sheet['f3'].value
book['pages'] = sheet['g3'].value
book['price'] = sheet['h3'].value
book['quantity'] = sheet['i3'].value
genre = sheet['j3'].value


print('---' * 10)
print('BOOK: ', book)
print('AUTHOR: ', author)
print('GENRE: ', genre)
